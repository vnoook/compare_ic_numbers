# ...
# INSTALL
# pip install openpyxl
# pip install PyQt5
# COMPILE
# pyinstaller -F file.py
# ...

import PyQt5
import PyQt5.QtWidgets
import sys
import openpyxl
import openpyxl.utils
import openpyxl.styles


# класс главного окна
class Window(PyQt5.QtWidgets.QMainWindow):
    # описание главного окна
    def __init__(self):
        super(Window, self).__init__()

        # переменные, атрибуты
        self.info_for_open_file = ''
        self.info_path_open_file = ''
        self.info_extention_open_file = 'Файлы Excel xlsx (*.xlsx)'
        self.text_empty_path_file = 'файл пока не выбран'
        self.text_empty_combobox = 'не выбрано'
        self.file_IC = ''
        self.file_GASPS = ''
        self.wb_file_IC = ''
        self.wb_file_IC_s = ''
        self.wb_file_GASPS = ''
        self.wb_file_GASPS_s = ''

        # главное окно, надпись на нём и размеры
        self.setWindowTitle('Сравнение номеров дел')
        self.setGeometry(300, 300, 900, 300)

        # объекты на главном окне
        # label_select_file_IC
        self.label_select_file_IC = PyQt5.QtWidgets.QLabel(self)
        self.label_select_file_IC.setObjectName('label_select_file_IC')
        self.label_select_file_IC.setText('1. Выберите файл ИЦ')
        self.label_select_file_IC.setGeometry(PyQt5.QtCore.QRect(10, 10, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_select_file_IC.setFont(font)
        self.label_select_file_IC.adjustSize()
        self.label_select_file_IC.setToolTip(self.label_select_file_IC.objectName())

        # toolButton_select_file_IC
        self.toolButton_select_file_IC = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_file_IC.setObjectName('toolButton_select_file_IC')
        self.toolButton_select_file_IC.setText('...')
        self.toolButton_select_file_IC.setGeometry(PyQt5.QtCore.QRect(10, 40, 50, 20))
        self.toolButton_select_file_IC.setFixedWidth(50)
        self.toolButton_select_file_IC.clicked.connect(self.select_file)
        self.toolButton_select_file_IC.setToolTip(self.toolButton_select_file_IC.objectName())

        # label_path_file_IC
        self.label_path_file_IC = PyQt5.QtWidgets.QLabel(self)
        self.label_path_file_IC.setObjectName('label_path_file_IC')
        self.label_path_file_IC.setText(self.text_empty_path_file)
        self.label_path_file_IC.setGeometry(PyQt5.QtCore.QRect(70, 40, 820, 16))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_file_IC.setFont(font)
        self.label_path_file_IC.adjustSize()
        self.label_path_file_IC.setToolTip(self.label_path_file_IC.objectName())

        # comboBox_liter_IC
        self.comboBox_liter_IC = PyQt5.QtWidgets.QComboBox(self)
        self.comboBox_liter_IC.setObjectName('comboBox_liter_IC')
        self.comboBox_liter_IC.setGeometry(PyQt5.QtCore.QRect(10, 70, 70, 20))
        self.comboBox_liter_IC.addItem('пусто')
        self.comboBox_liter_IC.setEnabled(False)
        self.comboBox_liter_IC.adjustSize()
        self.comboBox_liter_IC.setToolTip(self.comboBox_liter_IC.objectName())

        # comboBox_digit_IC
        self.comboBox_digit_IC = PyQt5.QtWidgets.QComboBox(self)
        self.comboBox_digit_IC.setObjectName('comboBox_digit_IC')
        self.comboBox_digit_IC.setGeometry(PyQt5.QtCore.QRect(110, 70, 70, 20))
        self.comboBox_digit_IC.addItem('пусто')
        self.comboBox_digit_IC.setEnabled(False)
        self.comboBox_digit_IC.adjustSize()
        self.comboBox_digit_IC.setToolTip(self.comboBox_digit_IC.objectName())

        # label_select_file_GASPS
        self.label_select_file_GASPS = PyQt5.QtWidgets.QLabel(self)
        self.label_select_file_GASPS.setObjectName('label_select_file_GASPS')
        self.label_select_file_GASPS.setText('2. Выберите файл ГАС ПС')
        self.label_select_file_GASPS.setGeometry(PyQt5.QtCore.QRect(10, 120, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_select_file_GASPS.setFont(font)
        self.label_select_file_GASPS.adjustSize()
        self.label_select_file_GASPS.setToolTip(self.label_select_file_GASPS.objectName())

        # label_path_file_GASPS
        self.label_path_file_GASPS = PyQt5.QtWidgets.QLabel(self)
        self.label_path_file_GASPS.setObjectName('label_path_file_GASPS')
        self.label_path_file_GASPS.setText(self.text_empty_path_file)
        self.label_path_file_GASPS.setGeometry(PyQt5.QtCore.QRect(70, 150, 820, 20))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_file_GASPS.setFont(font)
        self.label_path_file_GASPS.adjustSize()
        self.label_path_file_GASPS.setToolTip(self.label_path_file_GASPS.objectName())

        # toolButton_select_file_GASPS
        self.toolButton_select_file_GASPS = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_file_GASPS.setObjectName('toolButton_select_file_GASPS')
        self.toolButton_select_file_GASPS.setText('...')
        self.toolButton_select_file_GASPS.setGeometry(PyQt5.QtCore.QRect(10, 150, 50, 20))
        self.toolButton_select_file_GASPS.setFixedWidth(50)
        self.toolButton_select_file_GASPS.clicked.connect(self.select_file)
        self.toolButton_select_file_GASPS.setToolTip(self.toolButton_select_file_GASPS.objectName())

        # comboBox_liter_GASPS
        self.comboBox_liter_GASPS = PyQt5.QtWidgets.QComboBox(self)
        self.comboBox_liter_GASPS.setObjectName('comboBox_liter_GASPS')
        self.comboBox_liter_GASPS.setGeometry(PyQt5.QtCore.QRect(10, 180, 70, 20))
        self.comboBox_liter_GASPS.addItem('пусто')
        self.comboBox_liter_GASPS.setEnabled(False)
        self.comboBox_liter_GASPS.adjustSize()
        self.comboBox_liter_GASPS.setToolTip(self.comboBox_liter_GASPS.objectName())

        # comboBox_digit_GASPS
        self.comboBox_digit_GASPS = PyQt5.QtWidgets.QComboBox(self)
        self.comboBox_digit_GASPS.setObjectName('comboBox_digit_GASPS')
        self.comboBox_digit_GASPS.setGeometry(PyQt5.QtCore.QRect(110, 180, 70, 20))
        self.comboBox_digit_GASPS.addItem('пусто')
        self.comboBox_digit_GASPS.setEnabled(False)
        self.comboBox_digit_GASPS.adjustSize()
        self.comboBox_digit_GASPS.setToolTip(self.comboBox_digit_GASPS.objectName())

        # pushButton_do_fill_data
        self.pushButton_do_fill_data = PyQt5.QtWidgets.QPushButton(self)
        self.pushButton_do_fill_data.setObjectName('pushButton_do_fill_data')
        self.pushButton_do_fill_data.setEnabled(False)
        self.pushButton_do_fill_data.setText('Произвести заполнение')
        self.pushButton_do_fill_data.setGeometry(PyQt5.QtCore.QRect(10, 225, 180, 25))
        self.pushButton_do_fill_data.setFixedWidth(130)
        self.pushButton_do_fill_data.clicked.connect(self.do_fill_data)
        self.pushButton_do_fill_data.setToolTip(self.pushButton_do_fill_data.objectName())

        # кнопка button_exit
        self.button_exit = PyQt5.QtWidgets.QPushButton(self)
        self.button_exit.setObjectName('button_exit')
        self.button_exit.setText('Выход')
        self.button_exit.setGeometry(PyQt5.QtCore.QRect(10, 260, 180, 25))
        self.button_exit.setFixedWidth(50)
        self.button_exit.clicked.connect(self.click_on_btn_exit)
        self.button_exit.setToolTip(self.button_exit.objectName())

    # событие - нажатие на кнопку выбора файла
    def select_file(self):
        # запоминание старого значения пути выбора файлов
        old_path_of_selected_file_IC = self.label_path_file_IC.text()
        old_path_of_selected_file_GASPS = self.label_path_file_GASPS.text()

        # определение какая кнопка выбора файла нажата
        # если ИЦ, то выдать в окно про ИЦ
        if self.sender().objectName() == self.toolButton_select_file_IC.objectName():
            self.info_for_open_file = 'Выберите файл ИЦ формата Excel, версии старше 2007 года (.XLSX)'
        # если ГАСПС, то выдать в окно про ГАСПС
        elif self.sender().objectName() == self.toolButton_select_file_GASPS.objectName():
            self.info_for_open_file = 'Выберите файл ГАС ПС формата Excel, версии старше 2007 года (.XLSX)'

        # непосредственное окно выбора файла и переменная для хранения пути файла
        data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                             self.info_for_open_file,
                                                                             self.info_path_open_file,
                                                                             self.info_extention_open_file)
        # вычленение пути файла из data_of_open_file_name
        file_name = data_of_open_file_name[0]

        # выбор где и что менять исходя из выбора пользователя
        # нажата кнопка выбора ИЦ
        if self.sender().objectName() == self.toolButton_select_file_IC.objectName():
            if file_name == '':
                self.label_path_file_IC.setText(old_path_of_selected_file_IC)
                self.label_path_file_IC.adjustSize()
            else:
                old_path_of_selected_file_IC = self.label_path_file_IC.text()

                self.label_path_file_IC.setText(file_name)
                self.label_path_file_IC.adjustSize()

        # нажата кнопка выбора ГАСПС
        if self.sender().objectName() == self.toolButton_select_file_GASPS.objectName():
            if file_name == '':
                self.label_path_file_GASPS.setText(old_path_of_selected_file_GASPS)
                self.label_path_file_GASPS.adjustSize()
            else:
                old_path_of_selected_file_GASPS = self.label_path_file_GASPS.text()
                self.label_path_file_GASPS.setText(file_name)
                self.label_path_file_GASPS.adjustSize()

        # активация и деактивация объектов на форме зависящее от выбраны ли все файлы и они разные
        if self.label_path_file_IC.text() != self.label_path_file_GASPS.text():
            if self.text_empty_path_file not in (self.label_path_file_IC.text(), self.label_path_file_GASPS.text()):
                # self.pushButton_do_fill_data.setEnabled(True)
                self.comboBox_liter_IC.setEnabled(True)
                self.comboBox_digit_IC.setEnabled(True)
                self.comboBox_liter_GASPS.setEnabled(True)
                self.comboBox_digit_GASPS.setEnabled(True)
                self.do_fill_comboboxes()
        else:
            # self.pushButton_do_fill_data.setEnabled(False)
            self.comboBox_liter_IC.setEnabled(False)
            self.comboBox_digit_IC.setEnabled(False)
            self.comboBox_liter_GASPS.setEnabled(False)
            self.comboBox_digit_GASPS.setEnabled(False)

    # заполнение комбобоксов
    def do_fill_comboboxes(self):
        # присвоение файлов
        self.file_IC = self.label_path_file_IC.text()
        self.file_GASPS = self.label_path_file_GASPS.text()

        # открывается файл "приёмник", назначается активный лист, выбирается диапазон ячеек
        self.wb_file_IC = openpyxl.load_workbook(self.file_IC)
        self.wb_file_IC_s = self.wb_file_IC.active
        self.wb_file_GASPS = openpyxl.load_workbook(self.file_GASPS)  #, read_only=True)
        self.wb_file_GASPS_s = self.wb_file_GASPS.active

        # вычисление максимальных строк и колонок в выбранных файлах
        max_row_IC = self.wb_file_IC_s.max_row
        max_col_IC = self.wb_file_IC_s.max_column
        max_row_GASPS = self.wb_file_GASPS_s.max_row
        max_col_GASPS = self.wb_file_GASPS_s.max_column

        # очистка комбобоксов и заполнение их буквами и числами существующим в файлах
        self.comboBox_liter_IC.clear()
        self.comboBox_liter_IC.addItem(self.text_empty_combobox)
        self.comboBox_liter_IC.adjustSize()
        for col_IC in range(1, max_col_IC + 1):
            self.comboBox_liter_IC.addItem(openpyxl.utils.cell.coordinate_from_string(self.wb_file_IC_s.cell(1, col_IC).coordinate)[0])

        self.comboBox_digit_IC.clear()
        self.comboBox_digit_IC.addItem(self.text_empty_combobox)
        self.comboBox_digit_IC.adjustSize()
        for row_IC in range(1, max_row_IC + 1):
            self.comboBox_digit_IC.addItem(str(openpyxl.utils.cell.coordinate_from_string(self.wb_file_IC_s.cell(row_IC, 1).coordinate)[1]))

        self.comboBox_liter_GASPS.clear()
        self.comboBox_liter_GASPS.addItem(self.text_empty_combobox)
        self.comboBox_liter_GASPS.adjustSize()
        for col_GASPS in range(1, max_col_GASPS + 1):
            self.comboBox_liter_GASPS.addItem(openpyxl.utils.cell.coordinate_from_string(self.wb_file_GASPS_s.cell(1, col_GASPS).coordinate)[0])

        self.comboBox_digit_GASPS.clear()
        self.comboBox_digit_GASPS.addItem(self.text_empty_combobox)
        self.comboBox_digit_GASPS.adjustSize()
        for row_GASPS in range(1, max_row_GASPS + 1):
            self.comboBox_digit_GASPS.addItem(str(openpyxl.utils.cell.coordinate_from_string(self.wb_file_GASPS_s.cell(row_GASPS, 1).coordinate)[1]))

        # TODO
        # сделать разблокировку pushButton_do_fill_data
        self.pushButton_do_fill_data.setEnabled(True)

    # событие - нажатие на кнопку заполнения файла
    def do_fill_data(self):
        # определение множеств
        set_data_IC = set()
        set_data_GASPS = set()

        # проверка на то что все комбобоксы заполнены
        if self.text_empty_combobox not in (self.comboBox_liter_IC.itemText(self.comboBox_liter_IC.currentIndex()),
                                            self.comboBox_digit_IC.itemText(self.comboBox_digit_IC.currentIndex()),
                                            self.comboBox_liter_GASPS.itemText(self.comboBox_liter_GASPS.currentIndex()),
                                            self.comboBox_digit_GASPS.itemText(self.comboBox_digit_GASPS.currentIndex())):
            # формируются диапазоны для обработки данных в файлах из комбобоксов
            range_file_IC = self.comboBox_liter_IC.itemText(self.comboBox_liter_IC.currentIndex()) +\
                            self.comboBox_digit_IC.itemText(self.comboBox_digit_IC.currentIndex()) +\
                            ':' +\
                            self.comboBox_liter_IC.itemText(self.comboBox_liter_IC.currentIndex()) +\
                            self.comboBox_digit_IC.itemText(self.comboBox_digit_IC.count()-1)

            range_file_GASPS = self.comboBox_liter_GASPS.itemText(self.comboBox_liter_GASPS.currentIndex()) +\
                            self.comboBox_digit_GASPS.itemText(self.comboBox_digit_GASPS.currentIndex()) +\
                            ':' +\
                            self.comboBox_liter_GASPS.itemText(self.comboBox_liter_GASPS.currentIndex()) +\
                            self.comboBox_digit_GASPS.itemText(self.comboBox_digit_GASPS.count()-1)

            # сформированные диапазоны из выбранных комбобоксов
            wb_IC_cells_range = self.wb_file_IC_s[range_file_IC]
            wb_GASPS_cells_range = self.wb_file_GASPS_s[range_file_GASPS]

            # формирование множества из обработанных значений ячеек GASPS
            for row_in_range_GASPS in wb_GASPS_cells_range:
                for cell_in_row_GASPS in row_in_range_GASPS:
                    indexR_GASPS = wb_GASPS_cells_range.index(row_in_range_GASPS)
                    indexC_GASPS = row_in_range_GASPS.index(cell_in_row_GASPS)

                    if wb_GASPS_cells_range[indexR_GASPS][indexC_GASPS].value == None:
                        wb_GASPS_cell_value = 'None'
                    else:
                        wb_GASPS_cell_value = wb_GASPS_cells_range[indexR_GASPS][indexC_GASPS].value

                    for ikud in wb_GASPS_cell_value.split(";"):
                        set_data_GASPS.add(ikud.strip().replace('.', ''))

            for row_in_range_IC in wb_IC_cells_range:
                for cell_in_row_IC in row_in_range_IC:
                    indexR_IC = wb_IC_cells_range.index(row_in_range_IC)
                    indexC_IC = row_in_range_IC.index(cell_in_row_IC)

                    # получение координаты и значения ячейки IC
                    wb_IC_cell_coord = wb_IC_cells_range[indexR_IC][indexC_IC].coordinate
                    if wb_IC_cells_range[indexR_IC][indexC_IC].value == None:
                        wb_IC_cell_value = 'None'
                    else:
                        wb_IC_cell_value = wb_IC_cells_range[indexR_IC][indexC_IC].value

                    set_data_IC.clear()
                    for ikud in str(wb_IC_cell_value).split(";"):
                        set_data_IC.add(ikud.strip().replace('.', ''))

                    for ikud in str(wb_IC_cell_value).split(";"):
                        ikud_split = ikud.strip().replace('.', '').replace(' ', '')

                        if (ikud_split in set_data_GASPS) and (ikud_split in set_data_IC):
                            wb_IC_cells_range[indexR_IC][indexC_IC].fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        elif ikud_split not in set_data_GASPS:
                            wb_IC_cells_range[indexR_IC][indexC_IC].fill = openpyxl.styles.PatternFill(start_color='878787', end_color='878787', fill_type='solid')

            # сохраняю файлы и закрываю их
            self.wb_file_IC.save(self.file_IC)
            # self.wb_file_GASPS.save(self.file_GASPS)
            self.wb_file_IC.close()
            self.wb_file_GASPS.close()

            print('_'*20, ' файлы сохранены и закрыты ', '_'*20)
            print(self.file_IC)

            # очистка переменных от повторного использования
            del set_data_IC
            del set_data_GASPS

        else:
            print()
            print(f'выберите все поля')


    # событие - нажатие на кнопку Выход
    def click_on_btn_exit(self):
        exit()


# создание основного окна
def main_app():
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    app_window_main = Window()
    app_window_main.show()
    sys.exit(app.exec_())


# запуск основного окна
if __name__ == '__main__':
    main_app()
